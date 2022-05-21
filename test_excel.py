from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

#--追加

#for x in range(1,101):
#    for y in range(1,101):
#        ws.cell(row=x,column=y)

a = 'A'
b = 'B'
#for x in range(1,101):
#    for y in range(1,101):
#        d = a + str(x)
#        D = b + str(x)
#        ws[d] = x
#        ws[D] = x

for x in range(1,11):
    for y in range(1,11):
        ws.cell(row=x,column=y)


#d = 10
#c = a + str(d)
#ws[c] = 10




# Save the file
wb.save("sample.xlsx")
